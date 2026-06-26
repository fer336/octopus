"""
Servicio de exportación a Excel, CSV y PDF.
Convierte datos estructurados en archivos descargables.
"""

import csv
import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ExportService:
    """Exporta datos tabulares a formatos Excel, CSV y PDF."""

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def to_excel(data: list[dict[str, Any]], sheet_name: str = "Datos") -> bytes:
        """
        Genera un archivo Excel (.xlsx) desde una lista de diccionarios.

        Args:
            data: Lista de filas como diccionarios {columna: valor}.
            sheet_name: Nombre de la hoja del libro.

        Returns:
            Bytes del archivo .xlsx.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name max 31 chars

        if not data:
            ws.append(["Sin datos"])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        headers = list(data[0].keys())
        ws.append(headers)

        # Estilo de encabezados
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = ExportService.HEADER_FONT
            cell.fill = ExportService.HEADER_FILL
            cell.alignment = ExportService.HEADER_ALIGNMENT

        # Filas de datos
        for row_data in data:
            ws.append(list(row_data.values()))

        # Ajustar ancho de columnas automáticamente
        for col_idx, header in enumerate(headers, 1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else "A"
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        # Alineación: números a la derecha, texto a la izquierda
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.alignment = ExportService.CELL_ALIGNMENT
                else:
                    cell.alignment = ExportService.TEXT_ALIGNMENT

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def to_csv(data: list[dict[str, Any]]) -> str:
        """
        Genera un string CSV desde una lista de diccionarios.

        Args:
            data: Lista de filas como diccionarios {columna: valor}.

        Returns:
            Contenido CSV como string.
        """
        output = io.StringIO()
        if not data:
            return ""

        writer = csv.DictWriter(output, fieldnames=list(data[0].keys()))
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()

    @staticmethod
    def _prepare_filename(prefix: str, extension: str) -> str:
        """Genera un nombre de archivo con timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"


# Mapeo de content-type por formato
EXPORT_CONTENT_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "pdf": "application/pdf",
}
