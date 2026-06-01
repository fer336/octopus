"""
Servicio de Importación/Exportación Excel.
Maneja la carga masiva de productos.
"""
import io
import math
from datetime import date
from decimal import Decimal
from uuid import UUID

import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditLog
from app.models.brand import Brand
from app.models.category import Category
from app.models.product import Product
from app.models.product_lot import ProductLot
from app.models.supplier import Supplier
from app.services.brand_service import BrandService
from app.services.product_lot_service import ProductLotService
from app.schemas.excel_schemas import (
    ImportConfirmRequest,
    ImportConfirmResponse,
    ImportDetectResponse,
    ImportPreviewResponse,
    ProductImportRow,
)
from app.utils.brand_normalization import normalize_brand_name


# Maps internal field ids (used by the column mapper UI) to the canonical
# Spanish DataFrame column names that the parse loop expects.
FIELD_TO_EXCEL_KEY: dict[str, str] = {
    "code": "codigo",
    "description": "nombre",
    "list_price": "precio_lista",
    "supplier_code": "codigo_proveedor",
    "discounts": "bonificaciones",
    "bonificaciones": "bonificaciones",
    "extra_cost": "cargo_extra",
    "profit_margin": "ganancia",
    "iva": "iva",
    "iva_rate": "iva",
    "stock": "stock",
    "current_stock": "stock",
    "unit": "unidad",
    "units_per_pack": "unidades_x_pack",
    "expiration": "vencimiento",
    "expiration_date": "vencimiento",
    "category": "categoria",
    "supplier": "proveedor",
    "minimum_stock": "stock_minimo",
    "brand": "marca",
    "cost_price": "precio_costo",
    "details": "detalles",
}


def _clean_excel_cell(value: object) -> str | None:
    """Stringifies a cell value for sample rows; returns None for NaN/None."""
    if value is None:
        return None
    try:
        import math
        if isinstance(value, float) and math.isnan(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value)
    # Cap oversized values to avoid bloating the response
    return text[:200] if len(text) > 200 else text


class ExcelService:
    def __init__(self, db: AsyncSession):
        self.db = db

    def detect_columns(self, file_content: bytes) -> ImportDetectResponse:
        """
        Reads the first sheet of an Excel file and returns column headers plus
        up to 3 sample rows of stringified cell values.
        Does NOT require a DB session.
        """
        try:
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        except Exception as e:
            raise ValueError(f"Error al leer el archivo Excel: {str(e)}")

        columns = [str(col) for col in df.columns]
        sample_df = df.head(3)
        sample_rows: list[list] = [
            [_clean_excel_cell(cell) for cell in row]
            for row in sample_df.itertuples(index=False, name=None)
        ]

        return ImportDetectResponse(
            columns=columns,
            sample_rows=sample_rows,
            total_rows=len(df),
        )

    @staticmethod
    def _safe_decimal(value: object, default: Decimal = Decimal(0)) -> Decimal:
        """Convierte un valor de celda Excel a Decimal tolerando formatos regionales."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return default
        try:
            if isinstance(value, Decimal):
                return value
            if isinstance(value, int) and not isinstance(value, bool):
                return Decimal(value)
            if isinstance(value, float) and math.isfinite(value):
                return Decimal(str(value))

            s = str(value).strip()
            if not s or s.lower() in ('nan', 'none', '-', ''):
                return default
            # Remove currency symbols and non-breaking spaces
            s = s.replace('$', '').replace('\xa0', '').replace(' ', '')
            # Handle Argentine Excel imports.
            # Thousands separator = "." (between groups of 3 digits).
            # Decimal separator = ",".
            # Examples: "7.000" → 7000, "1.234,56" → 1234.56, "10.5" → 10.5.
            if ',' in s and '.' in s:
                s = s.replace('.', '').replace(',', '.')
            elif ',' in s:
                # Only comma present → treat as decimal separator
                s = s.replace(',', '.')
            elif '.' in s:
                # Only dot present → distinguish thousands from decimal.
                # If the last group after "." is exactly 3 digits → it's a
                # thousands separator (e.g. "7.000", "12.345").
                # Otherwise → it's a decimal point (e.g. "10.5", "7.5").
                parts = s.rsplit('.', 1)
                if len(parts) == 2 and len(parts[1]) == 3 and all(c.isdigit() for c in parts[1]):
                    s = s.replace('.', '')
                # else: keep dot, Decimal() handles it as decimal separator
            return Decimal(s)
        except Exception:
            return default

    def _parse_discounts(self, discounts_str: str) -> tuple[Decimal, Decimal, Decimal]:
        """Parsea el string de descuentos (ej: '10+5+2') en tres valores."""
        if not discounts_str or pd.isna(discounts_str):
            return Decimal(0), Decimal(0), Decimal(0)

        discounts_str = str(discounts_str).strip()
        if not discounts_str or discounts_str == '0':
            return Decimal(0), Decimal(0), Decimal(0)

        try:
            parts = discounts_str.replace(',', '.').split('+')
            discounts = []
            for part in parts[:3]:  # Solo tomar los primeros 3
                part = part.strip()
                if part:
                    discounts.append(Decimal(part))

            d1 = discounts[0] if len(discounts) > 0 else Decimal(0)
            d2 = discounts[1] if len(discounts) > 1 else Decimal(0)
            d3 = discounts[2] if len(discounts) > 2 else Decimal(0)

            return d1, d2, d3
        except (ValueError, IndexError):
            return Decimal(0), Decimal(0), Decimal(0)

    def _calculate_prices(self, list_price: Decimal, d1: Decimal, d2: Decimal, d3: Decimal,
                         extra_cost: Decimal, profit_margin: Decimal, iva_rate: Decimal) -> tuple[Decimal, Decimal, str]:
        """Calcula precio neto y final, igual que el modelo Product."""
        # Precio con bonificaciones (neto base)
        net_base = list_price * (1 - d1 / 100) * (1 - d2 / 100) * (1 - d3 / 100)

        # Aplicar cargo extra
        net_with_extra = net_base * (1 + extra_cost / 100)

        # Aplicar ganancia/utilidad
        net_with_profit = net_with_extra * (1 + profit_margin / 100)

        # Precio final con IVA
        sale = net_with_profit * (1 + iva_rate / 100)

        net_price = round(net_with_profit, 2)
        sale_price = round(sale, 2)

        # Formato de descuento para mostrar
        discounts = [d for d in [d1, d2, d3] if d > 0]
        discount_display = "+".join([str(int(d)) for d in discounts]) if discounts else None

        return net_price, sale_price, discount_display

    async def preview_import(
        self,
        business_id: UUID,
        file_content: bytes,
        column_mapping: dict[str, str] | None = None,
    ) -> ImportPreviewResponse:
        """
        Parsea el Excel y retorna un preview de los productos a importar.
        Incluye validaciones y cálculos de precios.

        Args:
            column_mapping: Optional mapping of {excel_header: field_id} produced
                by ColumnMapperModal. When provided, the DataFrame columns are
                renamed to canonical Spanish keys before the parse loop runs,
                so the legacy logic remains byte-identical.
        """
        try:
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        except Exception as e:
            raise ValueError(f"Error al leer el archivo Excel: {str(e)}")

        # Apply column mapping when provided: rename excel headers → canonical keys
        if column_mapping:
            rename_dict: dict[str, str] = {}
            for excel_col, field_id in column_mapping.items():
                canonical = FIELD_TO_EXCEL_KEY.get(field_id)
                if canonical:
                    rename_dict[excel_col] = canonical
            if rename_dict:
                df = df.rename(columns=rename_dict)

        # Validar columnas requeridas
        required_cols = ['codigo', 'nombre', 'precio_lista']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(missing_cols)}")

        # Obtain existing categories and suppliers scoped to this tenant
        # (fixes latent tenant leak — previously queried without business_id filter)
        categories_query = select(Category).where(
            Category.business_id == business_id,
            Category.deleted_at.is_(None),
        )
        categories_result = await self.db.execute(categories_query)
        categories = {cat.name.lower(): cat for cat in categories_result.scalars().all()}

        suppliers_query = select(Supplier).where(
            Supplier.business_id == business_id,
            Supplier.deleted_at.is_(None),
        )
        suppliers_result = await self.db.execute(suppliers_query)
        suppliers = {sup.name.lower(): sup for sup in suppliers_result.scalars().all()}

        brands: dict[str, Brand] = {}
        if "marca" in df.columns:
            brands_query = select(Brand).where(
                Brand.business_id == business_id,
                Brand.deleted_at.is_(None),
            )
            brands_result = await self.db.execute(brands_query)
            brands = {
                brand.normalized_name: brand for brand in brands_result.scalars().all()
            }

        rows: list[ProductImportRow] = []
        valid_count = 0
        error_count = 0
        new_count = 0
        existing_count = 0

        for index, row in df.iterrows():
            row_number = index + 2  # +2 porque Excel empieza en 1 y la fila 1 es header

            code = str(row.get('codigo', '')).strip()
            if not code:
                rows.append(ProductImportRow(
                    row_number=row_number,
                    code="",
                    description="",
                    has_errors=True,
                    error_message="El código es obligatorio"
                ))
                error_count += 1
                continue

            description = str(row.get('nombre', '')).strip()
            if not description:
                rows.append(ProductImportRow(
                    row_number=row_number,
                    code=code,
                    description="",
                    has_errors=True,
                    error_message="La descripción es obligatoria"
                ))
                error_count += 1
                continue

            # Parsear campos
            try:
                supplier_code = str(row.get('codigo_proveedor', '')).strip() if pd.notna(row.get('codigo_proveedor')) else None
                list_price = self._safe_decimal(row.get('precio_lista'))

                if list_price <= 0:
                    raise ValueError("El precio de lista debe ser mayor a 0")

                d1, d2, d3 = self._parse_discounts(row.get('bonificaciones', ''))
                extra_cost = self._safe_decimal(row.get('cargo_extra'))
                profit_margin = self._safe_decimal(row.get('ganancia'))
                iva_rate = self._safe_decimal(row.get('iva'), default=Decimal(21))
                current_stock_raw = row.get('stock')
                current_stock = int(self._safe_decimal(current_stock_raw)) if pd.notna(current_stock_raw) else 0

                # Leer unidad, pack qty y vencimiento
                unit = str(row.get('unidad', 'unidad')).strip().lower() if pd.notna(row.get('unidad')) else 'unidad'
                units_per_pack = None
                if pd.notna(row.get('unidades_x_pack')):
                    try:
                        units_per_pack = int(row.get('unidades_x_pack'))
                    except (ValueError, TypeError):
                        pass
                expiration_date = None
                if pd.notna(row.get('vencimiento')):
                    try:
                        exp_date = pd.to_datetime(row.get('vencimiento'))
                        expiration_date = exp_date.strftime('%Y-%m-%d')
                    except Exception:
                        pass

                # Calcular precios
                net_price, sale_price, discount_display = self._calculate_prices(
                    list_price, d1, d2, d3, extra_cost, profit_margin, iva_rate
                )

                # Resolve category: match by name (case-insensitive), flag new ones
                category_id = None
                category_name = None
                category_is_new = False
                if pd.notna(row.get('categoria')):
                    raw_cat = str(row.get('categoria')).strip()
                    if raw_cat:
                        cat_key = raw_cat.lower()
                        category = categories.get(cat_key)
                        if category:
                            category_id = category.id
                            category_name = category.name
                        else:
                            # Will be auto-created at confirm time
                            category_name = raw_cat
                            category_is_new = True

                # Resolve supplier: match by name (case-insensitive), flag new ones
                supplier_id = None
                supplier_name = None
                supplier_is_new = False
                if pd.notna(row.get('proveedor')):
                    raw_sup = str(row.get('proveedor')).strip()
                    if raw_sup:
                        sup_key = raw_sup.lower()
                        supplier = suppliers.get(sup_key)
                        if supplier:
                            supplier_id = supplier.id
                            supplier_name = supplier.name
                        else:
                            # Will be auto-created at confirm time
                            supplier_name = raw_sup
                            supplier_is_new = True

                # Resolve brand with normalization so FV, F.V. and fv collapse
                brand_id = None
                brand_name = None
                brand_is_new = False
                if pd.notna(row.get('marca')):
                    raw_brand = str(row.get('marca')).strip()
                    if raw_brand:
                        brand_key = normalize_brand_name(raw_brand)
                        brand = brands.get(brand_key)
                        if brand:
                            brand_id = brand.id
                            brand_name = brand.name
                        else:
                            brand_name = raw_brand
                            brand_is_new = True

                # Verificar si el producto ya existe
                existing_query = select(Product).where(
                    Product.code == code,
                    Product.business_id == business_id,
                    Product.deleted_at.is_(None)
                )
                existing_result = await self.db.execute(existing_query)
                existing_product = existing_result.scalar_one_or_none()

                is_new = existing_product is None
                existing_id = existing_product.id if existing_product else None

                if is_new:
                    new_count += 1
                else:
                    existing_count += 1

                rows.append(ProductImportRow(
                    row_number=row_number,
                    code=code,
                    supplier_code=supplier_code,
                    description=description,
                    category_id=category_id,
                    category_name=category_name,
                    category_is_new=category_is_new,
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    supplier_is_new=supplier_is_new,
                    brand_id=brand_id,
                    brand_name=brand_name,
                    brand_is_new=brand_is_new,
                    list_price=list_price,
                    discount_1=d1,
                    discount_2=d2,
                    discount_3=d3,
                    extra_cost=extra_cost,
                    profit_margin=profit_margin,
                    iva_rate=iva_rate,
                    current_stock=current_stock,
                    unit=unit,
                    units_per_pack=units_per_pack,
                    expiration_date=expiration_date,
                    net_price=net_price,
                    sale_price=sale_price,
                    discount_display=discount_display,
                    has_errors=False,
                    is_new=is_new,
                    existing_id=existing_id
                ))
                valid_count += 1

            except Exception as e:
                rows.append(ProductImportRow(
                    row_number=row_number,
                    code=code,
                    description=description,
                    has_errors=True,
                    error_message=f"Error al procesar: {str(e)}"
                ))
                error_count += 1

        return ImportPreviewResponse(
            total_rows=len(df),
            valid_rows=valid_count,
            rows_with_errors=error_count,
            new_products=new_count,
            existing_products=existing_count,
            rows=rows
        )

    async def confirm_import(
        self,
        business_id: UUID,
        request: ImportConfirmRequest,
        user_id: UUID | None = None,
    ) -> ImportConfirmResponse:
        """
        Confirma la importación de productos después del preview.
        Crea/actualiza productos en la base de datos.

        Auto-creates categories and suppliers that were flagged as new during
        preview (category_is_new / supplier_is_new). Lookups are case-insensitive
        and scoped to business_id to avoid cross-tenant collisions.

        Si se provee user_id, los lotes creados se atribuyen al usuario
        y se registra Auditoría (lot_operation:create).
        """
        created = 0
        updated = 0
        errors: list[str] = []

        # In-memory caches so we don't hit the DB twice for the same name
        # within the same import batch. Keys are lowercased names.
        category_cache: dict[str, Category] = {}
        supplier_cache: dict[str, Supplier] = {}
        brand_cache: dict[str, Brand] = {}
        brand_service = BrandService(self.db)

        async def _resolve_category(name: str) -> UUID | None:
            """Return category id, creating one if it doesn't exist."""
            if not name or not name.strip():
                return None
            key = name.strip().lower()
            if key in category_cache:
                return category_cache[key].id
            result = await self.db.execute(
                select(Category).where(
                    Category.business_id == business_id,
                    Category.deleted_at.is_(None),
                )
            )
            all_cats = {c.name.lower(): c for c in result.scalars().all()}
            if key in all_cats:
                category_cache[key] = all_cats[key]
                return all_cats[key].id
            # Create new
            new_cat = Category(name=name.strip(), business_id=business_id)
            self.db.add(new_cat)
            await self.db.flush()
            category_cache[key] = new_cat
            return new_cat.id

        async def _resolve_supplier(name: str) -> UUID | None:
            """Return supplier id, creating one if it doesn't exist."""
            if not name or not name.strip():
                return None
            key = name.strip().lower()
            if key in supplier_cache:
                return supplier_cache[key].id
            result = await self.db.execute(
                select(Supplier).where(
                    Supplier.business_id == business_id,
                    Supplier.deleted_at.is_(None),
                )
            )
            all_sups = {s.name.lower(): s for s in result.scalars().all()}
            if key in all_sups:
                supplier_cache[key] = all_sups[key]
                return all_sups[key].id
            # Create new
            new_sup = Supplier(name=name.strip(), business_id=business_id)
            self.db.add(new_sup)
            await self.db.flush()
            supplier_cache[key] = new_sup
            return new_sup.id

        async def _resolve_brand(name: str) -> Brand | None:
            """Return brand, creating one if it doesn't exist."""
            if not name or not name.strip():
                return None
            key = normalize_brand_name(name)
            if key in brand_cache:
                return brand_cache[key]
            brand = await brand_service.resolve_or_create(business_id, name)
            brand_cache[key] = brand
            return brand

        for row in request.rows:
            # Saltar filas con errores
            if row.has_errors:
                errors.append(f"Fila {row.row_number}: {row.error_message}")
                continue

            try:
                # Resolve category/supplier — auto-create new ones if flagged
                resolved_category_id = row.category_id
                if row.category_is_new and row.category_name:
                    resolved_category_id = await _resolve_category(row.category_name)
                elif row.category_name and row.category_id is None:
                    # Fallback: name was provided but id was not set (e.g. re-submitted row)
                    resolved_category_id = await _resolve_category(row.category_name)

                resolved_supplier_id = row.supplier_id
                if row.supplier_is_new and row.supplier_name:
                    resolved_supplier_id = await _resolve_supplier(row.supplier_name)
                elif row.supplier_name and row.supplier_id is None:
                    resolved_supplier_id = await _resolve_supplier(row.supplier_name)

                resolved_brand_id = row.brand_id
                resolved_brand_name = row.brand_name
                if row.brand_is_new and row.brand_name:
                    brand = await _resolve_brand(row.brand_name)
                    if brand:
                        resolved_brand_id = brand.id
                        resolved_brand_name = brand.name
                elif row.brand_name and row.brand_id is None:
                    brand = await _resolve_brand(row.brand_name)
                    if brand:
                        resolved_brand_id = brand.id
                        resolved_brand_name = brand.name

                if row.is_new:
                    # Crear nuevo producto
                    initial_stock = row.current_stock
                    new_product = Product(
                        business_id=business_id,
                        code=row.code,
                        supplier_code=row.supplier_code,
                        description=row.description,
                        category_id=resolved_category_id,
                        supplier_id=resolved_supplier_id,
                        brand_id=resolved_brand_id,
                        brand=resolved_brand_name,
                        list_price=row.list_price,
                        discount_1=row.discount_1,
                        discount_2=row.discount_2,
                        discount_3=row.discount_3,
                        extra_cost=row.extra_cost,
                        profit_margin=row.profit_margin,
                        iva_rate=row.iva_rate,
                        unit=row.unit,
                        units_per_pack=row.units_per_pack,
                        cost_price=Decimal(0),  # Se puede calcular después
                    )
                    new_product.calculate_prices()
                    self.db.add(new_product)

                    # Crear lote inicial si hay stock
                    if initial_stock > 0:
                        await self.db.flush()
                        lot = ProductLot(
                            product_id=new_product.id,
                            business_id=business_id,
                            code=f"IMP-{str(new_product.id)[:8]}",
                            quantity=initial_stock,
                            initial_quantity=initial_stock,
                            received_date=date.today(),
                            created_by=user_id,
                        )
                        self.db.add(lot)

                        # Flushear para obtener lot.id antes de auditoría
                        await self.db.flush()

                        # Registrar auditoría de creación de lote
                        if user_id:
                            audit = AuditLog(
                                user_id=user_id,
                                business_id=business_id,
                                action="create",
                                resource_type="lot_operation",
                                resource_id=lot.id,
                                details={
                                    "product_id": str(new_product.id),
                                    "quantity": initial_stock,
                                    "code": lot.code,
                                    "delta": initial_stock,
                                    "source": "excel_import",
                                },
                            )
                            self.db.add(audit)

                    created += 1
                else:
                    # Actualizar producto existente
                    if row.existing_id:
                        from sqlalchemy.orm import selectinload

                        query = select(Product).options(
                            selectinload(Product.lots)
                        ).where(Product.id == row.existing_id)
                        result = await self.db.execute(query)
                        existing_product = result.scalar_one_or_none()

                        if existing_product:
                            existing_product.code = row.code
                            existing_product.supplier_code = row.supplier_code
                            existing_product.description = row.description
                            existing_product.category_id = resolved_category_id
                            existing_product.supplier_id = resolved_supplier_id
                            existing_product.brand_id = resolved_brand_id
                            existing_product.brand = resolved_brand_name
                            existing_product.list_price = row.list_price
                            existing_product.discount_1 = row.discount_1
                            existing_product.discount_2 = row.discount_2
                            existing_product.discount_3 = row.discount_3
                            existing_product.extra_cost = row.extra_cost
                            existing_product.profit_margin = row.profit_margin
                            existing_product.iva_rate = row.iva_rate
                            existing_product.unit = row.unit
                            existing_product.units_per_pack = row.units_per_pack

                            # Ajustar stock via lotes (current_stock es @property)
                            new_stock_val = row.current_stock
                            current_stock_val = int(existing_product.current_stock)
                            stock_diff = new_stock_val - current_stock_val
                            if stock_diff > 0:
                                lot = ProductLot(
                                    product_id=existing_product.id,
                                    business_id=business_id,
                                    quantity=stock_diff,
                                    initial_quantity=stock_diff,
                                    received_date=date.today(),
                                    created_by=user_id,
                                )
                                self.db.add(lot)

                                # Flushear para obtener lot.id antes de auditoría
                                await self.db.flush()

                                # Registrar auditoría de creación de lote
                                if user_id:
                                    audit = AuditLog(
                                        user_id=user_id,
                                        business_id=business_id,
                                        action="create",
                                        resource_type="lot_operation",
                                        resource_id=lot.id,
                                        details={
                                            "product_id": str(existing_product.id),
                                            "quantity": stock_diff,
                                            "code": lot.code,
                                            "delta": stock_diff,
                                            "source": "excel_import",
                                        },
                                    )
                                    self.db.add(audit)
                            elif stock_diff < 0:
                                lot_service = ProductLotService(self.db)
                                try:
                                    await lot_service.fifo_consume(
                                        product_id=existing_product.id,
                                        business_id=business_id,
                                        quantity=abs(stock_diff),
                                        user_id=user_id,
                                        reason="Ajuste por importación Excel",
                                    )
                                except ValueError:
                                    errors.append(
                                        f"Fila {row.row_number}: Stock insuficiente para aplicar reducción"
                                    )

                            existing_product.calculate_prices()
                            updated += 1
                        else:
                            errors.append(f"Fila {row.row_number}: Producto existente no encontrado")
            except Exception as e:
                errors.append(f"Fila {row.row_number}: {str(e)}")

        await self.db.commit()

        return ImportConfirmResponse(
            created=created,
            updated=updated,
            errors=errors
        )

    async def import_products(self, business_id: UUID, file_content: bytes) -> dict:
        """
        Importa productos desde un archivo Excel.
        Retorna un resumen de la operación.
        """
        try:
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        except Exception as e:
            raise ValueError(f"Error al leer el archivo Excel: {str(e)}")

        # Normalizar columnas
        # Mapeo esperado:
        # codigo -> code
        # codigo_proveedor -> supplier_code
        # nombre -> description
        # stock -> current_stock
        # precio_lista -> list_price
        # bonificaciones -> discount_display (se debe parsear)
        # cargo_extra -> extra_cost

        column_map = {
            'codigo': 'code',
            'codigo_proveedor': 'supplier_code',
            'nombre': 'description',
            'stock': 'current_stock',
            'precio_lista': 'list_price',
            'bonificaciones': 'discounts',
            'cargo_extra': 'extra_cost'
        }

        # Validar columnas requeridas
        required_cols = ['codigo', 'nombre', 'precio_lista']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(missing_cols)}")

        summary = {
            "processed": 0,
            "created": 0,
            "updated": 0,
            "errors": []
        }

        for index, row in df.iterrows():
            summary["processed"] += 1
            try:
                code = str(row.get('codigo', '')).strip()
                if not code:
                    continue

                # Parsear descuentos
                discounts_str = str(row.get('bonificaciones', '0'))
                discounts = [Decimal(d.strip()) for d in discounts_str.split('+') if d.strip().replace('.', '', 1).isdigit()]

                d1 = discounts[0] if len(discounts) > 0 else Decimal(0)
                d2 = discounts[1] if len(discounts) > 1 else Decimal(0)
                d3 = discounts[2] if len(discounts) > 2 else Decimal(0)

                # Datos del producto
                product_data = {
                    "code": code,
                    "supplier_code": str(row.get('codigo_proveedor', '')).strip() if pd.notna(row.get('codigo_proveedor')) else None,
                    "description": str(row.get('nombre', '')),
                    "brand": str(row.get('marca', '')).strip() if pd.notna(row.get('marca')) else None,
                    "current_stock": int(self._safe_decimal(row.get('stock'))) if pd.notna(row.get('stock')) else 0,
                    "list_price": self._safe_decimal(row.get('precio_lista')),
                    "discount_1": d1,
                    "discount_2": d2,
                    "discount_3": d3,
                    "extra_cost": self._safe_decimal(row.get('cargo_extra')),
                    "profit_margin": self._safe_decimal(row.get('ganancia')),
                    "iva_rate": Decimal("21.00"), # Default
                    "unit": str(row.get('unidad', 'unidad')).strip().lower() if pd.notna(row.get('unidad')) else 'unidad',
                    "units_per_pack": int(row.get('unidades_x_pack')) if pd.notna(row.get('unidades_x_pack')) else None,
                    "expiration_date": None,
                    "business_id": business_id
                }

                if product_data["brand"]:
                    brand = await BrandService(self.db).resolve_or_create(
                        business_id,
                        str(product_data["brand"]),
                    )
                    product_data["brand_id"] = brand.id
                    product_data["brand"] = brand.name

                # Parsear vencimiento si está presente
                if pd.notna(row.get('vencimiento')):
                    try:
                        exp_date = pd.to_datetime(row.get('vencimiento'))
                        product_data["expiration_date"] = exp_date.date()
                    except Exception:
                        pass

                # Buscar si existe
                query = select(Product).where(
                    Product.code == code,
                    Product.business_id == business_id,
                    Product.deleted_at.is_(None)
                )
                result = await self.db.execute(query)
                existing_product = result.scalar_one_or_none()

                if existing_product:
                    # Actualizar
                    for key, value in product_data.items():
                        if key != "business_id":
                            setattr(existing_product, key, value)

                    existing_product.calculate_prices()
                    summary["updated"] += 1
                else:
                    # Crear
                    new_product = Product(**product_data)
                    new_product.calculate_prices()
                    self.db.add(new_product)
                    summary["created"] += 1

            except Exception as e:
                summary["errors"].append(f"Fila {index + 2}: {str(e)}")

        await self.db.commit()
        return summary

    async def export_products(
        self,
        business_id: UUID,
        category_id: UUID | None = None,
        supplier_id: UUID | None = None,
    ) -> bytes:
        """
        Exporta productos a Excel con formato profesional.
        Solo incluye las columnas esenciales para importación.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from sqlalchemy.orm import selectinload

        query = (
            select(Product)
            .options(
                selectinload(Product.category),
                selectinload(Product.supplier),
                selectinload(Product.brand_ref),
                selectinload(Product.lots),
            )
            .where(
                Product.business_id == business_id,
                Product.deleted_at.is_(None),
            )
            .order_by(Product.code)
        )

        if category_id is not None:
            query = query.where(Product.category_id == category_id)
        if supplier_id is not None:
            query = query.where(Product.supplier_id == supplier_id)

        result = await self.db.execute(query)
        products = result.scalars().all()

        # Crear workbook manualmente para control total del formato
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Definir columnas (orden para importación)
        headers = [
            'codigo', 'codigo_proveedor', 'nombre_proveedor', 'categoria',
            'nombre', 'marca', 'unidad', 'stock', 'precio_lista', 'bonificaciones',
            'cargo_extra', 'ganancia', 'vencimiento', 'unidades_x_pack', 'cantidad_por_compra', 'precio_venta'
        ]

        # Escribir headers con estilo
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Escribir datos
        for row_idx, p in enumerate(products, 2):
            # Formato de bonificaciones (ej: "10+5+2")
            discounts = [
                float(p.discount_1 or 0),
                float(p.discount_2 or 0),
                float(p.discount_3 or 0)
            ]
            bonificaciones_str = '+'.join([str(int(d)) for d in discounts if d > 0]) or '0'

            row_data = [
                p.code or '',
                p.supplier_code or '',
                p.supplier.name if p.supplier else '',
                p.category.name if p.category else '',
                p.description or '',
                p.brand_ref.name if p.brand_ref else (p.brand or ''),
                p.unit or 'unidad',
                int(p.current_stock) if p.current_stock else 0,
                float(p.list_price) if p.list_price else 0.0,
                bonificaciones_str,
                float(p.extra_cost) if p.extra_cost else 0.0,
                float(p.profit_margin) if p.profit_margin else 0.0,
                p.next_expiration.strftime('%Y-%m-%d') if p.next_expiration else '',
                int(p.units_per_pack) if p.units_per_pack else '',
                float(p.quantity_per_package) if p.quantity_per_package else '',
                float(p.sale_price) if p.sale_price else 0.0,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Alineación según tipo
                if col_idx in [3, 4, 5]:  # nombre_proveedor, categoria, nombre
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx in [1, 2, 6, 9, 12, 13, 14, 15]:  # códigos, unidad, bonificaciones, vencimiento, unidades_x_pack, cantidad_por_compra
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:  # números
                    cell.alignment = Alignment(horizontal="right", vertical="center")

                # Formato de números
                if col_idx in [8, 10, 11, 16]:  # precio_lista, cargo_extra, ganancia, precio_venta
                    cell.number_format = '#,##0.00'
                elif col_idx == 7:  # stock
                    cell.number_format = '0'

                # Bordes
                cell.border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )

                # Alternar color de filas
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Ajustar anchos de columna
        column_widths = {
            'A': 15,  # codigo
            'B': 18,  # codigo_proveedor
            'C': 25,  # nombre_proveedor
            'D': 20,  # categoria
            'E': 40,  # nombre (descripción del producto)
            'F': 10,  # unidad
            'G': 12,  # stock
            'H': 15,  # precio_lista
            'I': 16,  # bonificaciones
            'J': 14,  # cargo_extra
            'K': 14,  # ganancia
            'L': 15,  # vencimiento
            'M': 16,  # unidades_x_pack
            'N': 16,  # cantidad_por_compra
            'O': 15,  # precio_venta
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze primera fila (header)
        ws.freeze_panes = "A2"

        # Crear hoja de REFERENCIA con categorías y proveedores
        ws_ref = wb.create_sheet(title="Referencia")

        # Header de referencia
        ws_ref['A1'] = 'CATEGORÍAS DISPONIBLES'
        ws_ref['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_ref['A1'].fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        ws_ref['C1'] = 'PROVEEDORES DISPONIBLES'
        ws_ref['C1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws_ref['C1'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

        # Obtener categorías y proveedores para la referencia
        categories_query = select(Category).where(
            Category.business_id == business_id,
            Category.deleted_at.is_(None)
        ).order_by(Category.name)

        categories_result = await self.db.execute(categories_query)
        categories = categories_result.scalars().all()

        suppliers_query = select(Supplier).where(
            Supplier.business_id == business_id,
            Supplier.deleted_at.is_(None)
        ).order_by(Supplier.name)

        suppliers_result = await self.db.execute(suppliers_query)
        suppliers = suppliers_result.scalars().all()

        # Escribir categorías
        for idx, cat in enumerate(categories, 2):
            cell = ws_ref.cell(row=idx, column=1, value=cat.name)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        # Escribir proveedores
        for idx, sup in enumerate(suppliers, 2):
            cell = ws_ref.cell(row=idx, column=3, value=sup.name)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

        # Ajustar anchos de la hoja de referencia
        ws_ref.column_dimensions['A'].width = 30
        ws_ref.column_dimensions['C'].width = 30

        # Instrucciones
        ws_ref['A' + str(len(categories) + 4)] = '💡 Copia estos nombres exactos a tu Excel en las columnas categoria/proveedor'
        ws_ref['A' + str(len(categories) + 4)].font = Font(italic=True, color="666666")

        # Guardar a BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    async def export_full_backup(self, business_id: UUID) -> bytes:
        """
        Exporta TODOS los productos (incluso eliminados) como backup completo de la DB.
        Incluye TODOS los campos para poder restaurar.
        """
        from sqlalchemy.orm import selectinload

        # Obtener TODOS los productos, incluso los eliminados
        query = (
            select(Product)
            .options(
                selectinload(Product.category),
                selectinload(Product.supplier),
                selectinload(Product.lots),
            )
            .where(Product.business_id == business_id)
            .order_by(Product.created_at)
        )

        result = await self.db.execute(query)
        products = result.scalars().all()

        data = []
        for p in products:
            data.append({
                'id': str(p.id),
                'codigo': p.code,
                'codigo_proveedor': p.supplier_code or '',
                'nombre': p.description,
                'marca': p.brand or '',
                'detalles': p.details or '',
                'categoria': p.category.name if p.category else '',
                'categoria_id': str(p.category_id) if p.category_id else '',
                'proveedor': p.supplier.name if p.supplier else '',
                'proveedor_id': str(p.supplier_id) if p.supplier_id else '',
                'precio_costo': float(p.cost_price),
                'precio_lista': float(p.list_price),
                'descuento_1': float(p.discount_1),
                'descuento_2': float(p.discount_2),
                'descuento_3': float(p.discount_3),
                'bonificaciones': p.discount_display or "0",
                'cargo_extra': float(p.extra_cost),
                'ganancia': float(p.profit_margin),
                'precio_neto': float(p.net_price),
                'precio_venta': float(p.sale_price),
                'iva': float(p.iva_rate),
                'stock_actual': p.current_stock,
                'stock_minimo': p.minimum_stock,
                'unidad': p.unit,
                'unidades_x_pack': p.units_per_pack,
                'cantidad_por_compra': float(p.quantity_per_package) if p.quantity_per_package else '',
                'vencimiento': p.expiration_date.strftime('%Y-%m-%d') if p.expiration_date else '',
                'activo': p.is_active,
                'fecha_creacion': p.created_at.isoformat() if p.created_at else '',
                'fecha_actualizacion': p.updated_at.isoformat() if p.updated_at else '',
                'fecha_eliminacion': p.deleted_at.isoformat() if p.deleted_at else '',
            })

        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Backup_Productos')

        return output.getvalue()

    async def delete_all_products(self, business_id: UUID) -> dict:
        """
        Elimina TODOS los productos de un negocio (soft delete).
        Retorna un resumen de la operación.
        """
        from datetime import datetime

        query = select(Product).where(
            Product.business_id == business_id,
            Product.deleted_at.is_(None)
        )

        result = await self.db.execute(query)
        products = result.scalars().all()

        count = 0
        now = datetime.utcnow()

        for product in products:
            product.deleted_at = now
            count += 1

        await self.db.commit()

        return {
            "deleted_count": count,
            "message": f"Se eliminaron {count} productos correctamente"
        }
