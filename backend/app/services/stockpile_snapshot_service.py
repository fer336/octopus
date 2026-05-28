"""
Servicio de Snapshots de Precios de Acopio.
Generación de Excel de precios congelados.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from decimal import Decimal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.stockpile import StockpilePriceSnapshot

logger = logging.getLogger("uvicorn")

# Estilos para el Excel
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
MONEY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00"%"'
DATE_FORMAT = 'DD/MM/YYYY HH:MM'


class StockpileSnapshotService:
    """Servicio para manejar snapshots de precios de acopios."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_snapshots(
        self, stockpile_id: UUID
    ) -> list[StockpilePriceSnapshot]:
        """Obtiene todos los snapshots de un acopio."""
        result = await self.db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile_id,
                StockpilePriceSnapshot.deleted_at.is_(None),
            ).order_by(StockpilePriceSnapshot.code)
        )
        return list(result.scalars().all())

    def generate_excel(
        self,
        snapshots: list[StockpilePriceSnapshot],
        stockpile_name: str,
    ) -> io.BytesIO:
        """
        Genera un archivo Excel con los snapshots de precios.

        Columnas: Código, Descripción, Precio Final
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Precios Congelados"

        # Título
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = f"Precios congelados - {stockpile_name}"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="1f2937")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Headers
        headers = [
            "Código",
            "Descripción",
            "Precio Final",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        ws.row_dimensions[3].height = 25

        # Datos
        for row_num, snap in enumerate(snapshots, 4):
            ws.cell(
                row=row_num, column=1, value=snap.code
            ).alignment = LEFT_ALIGNMENT

            ws.cell(
                row=row_num, column=2, value=snap.description
            ).alignment = LEFT_ALIGNMENT

            total_cell = ws.cell(
                row=row_num, column=3, value=float(snap.price_with_iva)
            )
            total_cell.number_format = MONEY_FORMAT
            total_cell.alignment = CELL_ALIGNMENT

        # Ajustar ancho de columnas
        column_widths = [15, 50, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Autofilter en el rango de datos
        if snapshots:
            ws.auto_filter.ref = f"A3:C{3 + len(snapshots)}"

        # Freeze panes: fila de headers + columna código
        ws.freeze_panes = "B4"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
