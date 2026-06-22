"""
Unit tests for price list export services (Excel and PDF).
No DB access — uses SimpleNamespace mocks.
"""

import types
from datetime import date
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock

import pytest


# ---------------------------------------------------------------------------
# Mock helpers
# ---------------------------------------------------------------------------


def _make_price_list(name: str = "Test List") -> types.SimpleNamespace:
    return types.SimpleNamespace(
        id="00000000-0000-0000-0000-000000000001",
        name=name,
        valid_from=date(2025, 1, 1),
        valid_until=date(2025, 12, 31),
        currency="ARS",
        includes_tax=True,
        status="active",
        description=None,
        terms_and_conditions=None,
        notes=None,
        items=[],
    )


def _make_item(
    product_code: str = "PROD001",
    unit_price: Decimal = Decimal("121.00"),
) -> types.SimpleNamespace:
    return types.SimpleNamespace(
        product_code=product_code,
        supplier_code="SUP001",
        brand_name="BrandX",
        category_name="Electronics",
        description=None,
        unit="unidad",
        pack_quantity=Decimal("12"),
        min_quantity=Decimal("1"),
        base_price=Decimal("100.00"),
        discount_percent=Decimal("0"),
        surcharge_percent=Decimal("0"),
        net_price=Decimal("100.00"),
        iva_rate=Decimal("21"),
        tax_percent=Decimal("21"),
        final_price=Decimal("121.00"),
        unit_price=unit_price,
        item_notes=None,
    )


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_excel_export_bytes():
    """export_price_list returns non-empty bytes."""
    from app.services.excel_service import ExcelService

    mock_db = AsyncMock()
    svc = ExcelService(mock_db)

    pl = _make_price_list()
    items = [_make_item("A001"), _make_item("A002")]

    result = await svc.export_price_list(pl, items)

    assert isinstance(result, bytes)
    assert len(result) > 0


@pytest.mark.asyncio
async def test_excel_export_bytes_empty_items():
    """export_price_list works with an empty items list."""
    from app.services.excel_service import ExcelService

    mock_db = AsyncMock()
    svc = ExcelService(mock_db)

    pl = _make_price_list("Empty List")

    result = await svc.export_price_list(pl, [])

    assert isinstance(result, bytes)
    assert len(result) > 0


@pytest.mark.asyncio
async def test_excel_no_cost_fields():
    """Excel headers must not include internal cost or margin columns."""
    import io

    from openpyxl import load_workbook

    from app.services.excel_service import ExcelService

    mock_db = AsyncMock()
    svc = ExcelService(mock_db)

    pl = _make_price_list()
    items = [_make_item()]

    raw = await svc.export_price_list(pl, items)

    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active

    # Collect all header cell values (row 2 is the header row)
    headers_lower = [
        str(ws.cell(row=2, column=c).value or "").lower()
        for c in range(1, ws.max_column + 1)
    ]

    forbidden = {"costo", "margen", "ganancia", "profit", "cost", "margin"}
    found = {h for h in headers_lower if any(f in h for f in forbidden)}
    assert not found, f"Forbidden cost/margin headers found: {found}"


@pytest.mark.asyncio
async def test_excel_export_prefers_product_description_over_snapshot():
    """Excel export must show the linked product name, not a stale snapshot category."""
    import io

    from openpyxl import load_workbook

    from app.services.excel_service import ExcelService

    mock_db = AsyncMock()
    svc = ExcelService(mock_db)

    pl = _make_price_list()
    item = _make_item()
    item.description = "Pintureria"
    item.product = types.SimpleNamespace(description="Rodillo profesional")

    raw = await svc.export_price_list(pl, [item])
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active

    assert ws.cell(row=3, column=5).value == "Rodillo profesional"


# ---------------------------------------------------------------------------
# PDF export tests
# ---------------------------------------------------------------------------


def test_pdf_generates_bytes():
    """generate_price_list_pdf returns valid PDF bytes (starts with %PDF)."""
    weasyprint = pytest.importorskip("weasyprint", reason="WeasyPrint not installed")

    from app.services.pdf_service import pdf_service

    business = types.SimpleNamespace(
        name="Acme Corp",
        address="123 Main St",
        cuit="30-12345678-9",
    )
    pl = _make_price_list("Q1 B2B")
    context = {
        "business": business,
        "price_list": pl,
        "items": [_make_item()],
        "issued_date": date.today(),
    }

    result = pdf_service.generate_price_list_pdf(context)

    assert isinstance(result, bytes)
    assert result[:4] == b"%PDF", "Result is not a valid PDF"
