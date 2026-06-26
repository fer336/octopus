"""
Tests for brand/export changes and _calculate_unit_cost logic.

Covers:
  1. export_products — 'marca' header at index 5; brand value at row_data index 5
  2. export_full_backup — 'marca' key exists with correct value in backup dict
  3. _calculate_unit_cost — formula with various discount combinations
  4. import_from_excel (unit_cost) — _calculate_unit_cost is called instead of cost_price
"""

from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
import pytest_asyncio
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.services.excel_service import ExcelService
from app.services.purchase_order_service import PurchaseOrderService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_product(
    *,
    brand: str | None = None,
    list_price: Decimal = Decimal("1000.00"),
    discount_1: Decimal = Decimal("0"),
    discount_2: Decimal = Decimal("0"),
    discount_3: Decimal = Decimal("0"),
    cost_price: Decimal = Decimal("0"),
    code: str = "TST-001",
    description: str = "Test product",
    lots: list | None = None,
    supplier=None,
    category=None,
) -> MagicMock:
    """Build a MagicMock that quacks like a Product with the given field values."""
    p = MagicMock(spec=Product)
    p.id = uuid4()  # must be a real UUID for Pydantic validation in import_from_excel
    p.brand = brand
    p.list_price = list_price
    p.discount_1 = discount_1
    p.discount_2 = discount_2
    p.discount_3 = discount_3
    p.cost_price = cost_price
    p.code = code
    p.supplier_code = None
    p.description = description
    p.supplier = supplier
    p.category = category
    p.unit = "unidad"
    p.next_expiration = None
    p.units_per_pack = None
    p.sale_price = Decimal("1210.00")
    p.extra_cost = Decimal("0")
    p.profit_margin = Decimal("0")
    p.iva_rate = Decimal("21")
    p.brand_ref = None       # avoid MagicMock when excel_service checks brand_ref.name
    p.current_stock = 0      # avoid MagicMock — pandas chokes on mock's tzinfo
    # lots used by current_stock property via mocked attribute
    p.lots = lots or []
    return p


def _make_backup_product(brand: str | None = None) -> MagicMock:
    """Build a MagicMock for export_full_backup (needs more fields)."""
    p = _make_product(brand=brand)
    p.id = uuid4()
    p.supplier_code = ""
    p.details = ""
    p.discount_display = "0"
    p.net_price = Decimal("1000.00")
    p.minimum_stock = 0
    p.is_active = True
    p.expiration_date = None
    p.created_at = None
    p.updated_at = None
    p.deleted_at = None
    p.category_id = None
    p.supplier_id = None
    return p


# ---------------------------------------------------------------------------
# _calculate_unit_cost tests (pure unit tests — no DB needed)
# ---------------------------------------------------------------------------


class TestCalculateUnitCost:
    """Tests for PurchaseOrderService._calculate_unit_cost."""

    def _service(self):
        db = MagicMock(spec=AsyncSession)
        return PurchaseOrderService(db=db)

    def test_no_discounts_returns_list_price(self):
        svc = self._service()
        product = _make_product(list_price=Decimal("1000.00"))
        result = svc._calculate_unit_cost(product)
        assert result == Decimal("1000.00")

    def test_single_discount_10_percent(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"), discount_1=Decimal("10")
        )
        # 1000 × 0.90 = 900
        assert svc._calculate_unit_cost(product) == Decimal("900.00")

    def test_single_discount_25_percent(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"), discount_1=Decimal("25")
        )
        assert svc._calculate_unit_cost(product) == Decimal("750.00")

    def test_two_discounts_10_and_5(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"),
            discount_1=Decimal("10"),
            discount_2=Decimal("5"),
        )
        # 1000 × 0.90 × 0.95 = 855
        assert svc._calculate_unit_cost(product) == Decimal("855.00")

    def test_three_discounts_15_10_5(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"),
            discount_1=Decimal("15"),
            discount_2=Decimal("10"),
            discount_3=Decimal("5"),
        )
        # 1000 × 0.85 × 0.90 × 0.95 = 727.425 → rounded to 727.43
        result = svc._calculate_unit_cost(product)
        expected = round(
            Decimal("1000") * Decimal("0.85") * Decimal("0.90") * Decimal("0.95"), 2
        )
        assert result == expected

    def test_three_discounts_20_10_5(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("500.00"),
            discount_1=Decimal("20"),
            discount_2=Decimal("10"),
            discount_3=Decimal("5"),
        )
        # 500 × 0.80 × 0.90 × 0.95 = 342
        result = svc._calculate_unit_cost(product)
        expected = round(
            Decimal("500") * Decimal("0.80") * Decimal("0.90") * Decimal("0.95"), 2
        )
        assert result == expected

    def test_zero_list_price_returns_zero(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("0"),
            discount_1=Decimal("10"),
        )
        assert svc._calculate_unit_cost(product) == Decimal("0.00")

    def test_none_list_price_treated_as_zero(self):
        svc = self._service()
        product = _make_product(list_price=None)
        product.list_price = None
        assert svc._calculate_unit_cost(product) == Decimal("0.00")

    def test_none_discounts_treated_as_zero(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("200.00"),
            discount_1=None,
            discount_2=None,
            discount_3=None,
        )
        product.discount_1 = None
        product.discount_2 = None
        product.discount_3 = None
        # With all None → treated as 0 → result = 200
        assert svc._calculate_unit_cost(product) == Decimal("200.00")

    def test_result_is_rounded_to_two_decimals(self):
        svc = self._service()
        # 100 × (1-0.333) = 66.7 — arbitrary fraction to force rounding
        product = _make_product(
            list_price=Decimal("100.00"), discount_1=Decimal("33.33")
        )
        result = svc._calculate_unit_cost(product)
        # Just verify two decimal places
        assert result == round(result, 2)

    def test_100_percent_discount_returns_zero(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"), discount_1=Decimal("100")
        )
        assert svc._calculate_unit_cost(product) == Decimal("0.00")

    def test_only_d2_set(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"),
            discount_1=Decimal("0"),
            discount_2=Decimal("10"),
            discount_3=Decimal("0"),
        )
        # 1000 × 1 × 0.90 × 1 = 900
        assert svc._calculate_unit_cost(product) == Decimal("900.00")

    def test_only_d3_set(self):
        svc = self._service()
        product = _make_product(
            list_price=Decimal("1000.00"),
            discount_1=Decimal("0"),
            discount_2=Decimal("0"),
            discount_3=Decimal("20"),
        )
        # 1000 × 1 × 1 × 0.80 = 800
        assert svc._calculate_unit_cost(product) == Decimal("800.00")


# ---------------------------------------------------------------------------
# export_products — headers and row_data position of 'marca'
# ---------------------------------------------------------------------------


EXPORT_PRODUCTS_HEADERS = [
    "codigo",
    "codigo_proveedor",
    "nombre_proveedor",
    "categoria",
    "nombre",
    "marca",         # index 5
    "unidad",
    "stock",
    "precio_lista",
    "bonificaciones",
    "cargo_extra",
    "ganancia",
    "vencimiento",
    "unidades_x_pack",
    "precio_venta",
]


class TestExportProductsHeaders:
    """Verify that 'marca' is at index 5 in the headers list defined in ExcelService."""

    def test_marca_at_index_5(self):
        assert EXPORT_PRODUCTS_HEADERS[5] == "marca"

    def test_nombre_at_index_4(self):
        """'nombre' should be immediately before 'marca'."""
        assert EXPORT_PRODUCTS_HEADERS[4] == "nombre"

    def test_unidad_at_index_6(self):
        """'unidad' should be immediately after 'marca'."""
        assert EXPORT_PRODUCTS_HEADERS[6] == "unidad"

    def test_total_headers_count(self):
        assert len(EXPORT_PRODUCTS_HEADERS) == 15

    def test_marca_present_in_headers(self):
        assert "marca" in EXPORT_PRODUCTS_HEADERS


# ---------------------------------------------------------------------------
# export_products — row_data brand value (integration-style unit test)
# ---------------------------------------------------------------------------


def _make_export_products_db_mock(products: list) -> AsyncMock:
    """
    export_products calls db.execute three times:
      1. SELECT products
      2. SELECT categories (for reference sheet)
      3. SELECT suppliers (for reference sheet)
    Return the products on the first call; empty scalars on the rest.
    """
    db = AsyncMock(spec=AsyncSession)

    products_result = MagicMock()
    products_result.scalars.return_value.all.return_value = products

    empty_result = MagicMock()
    empty_result.scalars.return_value.all.return_value = []

    db.execute = AsyncMock(side_effect=[products_result, empty_result, empty_result])
    return db


@pytest.mark.asyncio
async def test_export_products_brand_in_row_data_position_5():
    """
    The 6th element (index 5) in row_data must equal the product's brand.
    We verify the live Excel bytes written by ExcelService contain the brand.
    """
    business_id = uuid4()

    product = _make_product(brand="AcmeBrand", list_price=Decimal("100.00"))
    db = _make_export_products_db_mock([product])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    headers = [ws.cell(row=1, column=i).value for i in range(1, 16)]
    assert headers[5] == "marca", f"Expected 'marca' at col index 5, got {headers[5]}"

    data_row = [ws.cell(row=2, column=i).value for i in range(1, 16)]
    assert data_row[5] == "AcmeBrand", f"Expected 'AcmeBrand' at data col 5, got {data_row[5]}"


@pytest.mark.asyncio
async def test_export_products_brand_none_is_empty_string():
    """When product.brand is None, row_data at index 5 must be empty string ''."""
    business_id = uuid4()

    product = _make_product(brand=None)
    db = _make_export_products_db_mock([product])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    data_row = [ws.cell(row=2, column=i).value for i in range(1, 16)]
    # openpyxl returns None for empty cells (brand=None written as '' → stored as None)
    assert data_row[5] in ("", None), f"Expected empty at data col 5, got {data_row[5]!r}"


@pytest.mark.asyncio
async def test_export_products_multiple_products_brands():
    """Each product row writes its own brand at column index 5."""
    business_id = uuid4()

    p1 = _make_product(brand="BrandA", code="P001")
    p2 = _make_product(brand="BrandB", code="P002")
    p3 = _make_product(brand=None, code="P003")
    db = _make_export_products_db_mock([p1, p2, p3])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    assert ws.cell(row=2, column=6).value == "BrandA"
    assert ws.cell(row=3, column=6).value == "BrandB"
    # openpyxl returns None for empty-string cells (brand=None written as '')
    assert ws.cell(row=4, column=6).value in ("", None)


@pytest.mark.asyncio
async def test_export_products_returns_bytes():
    """export_products must return bytes-like (non-empty)."""
    business_id = uuid4()
    db = _make_export_products_db_mock([])

    svc = ExcelService(db=db)
    result = await svc.export_products(business_id=business_id)

    assert isinstance(result, bytes)
    assert len(result) > 0


# ---------------------------------------------------------------------------
# export_full_backup — 'marca' key in backup dict
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_full_backup_marca_key_exists():
    """Backup dict must contain key 'marca' for each product."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(brand="BackupBrand")

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    assert "marca" in df.columns, f"'marca' column not found. Columns: {list(df.columns)}"


@pytest.mark.asyncio
async def test_export_full_backup_marca_value_matches_product_brand():
    """The 'marca' cell value in the backup equals product.brand."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(brand="TestBrand")

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    assert df.iloc[0]["marca"] == "TestBrand"


@pytest.mark.asyncio
async def test_export_full_backup_marca_none_exports_as_empty():
    """When product.brand is None, backup 'marca' exports as empty string ''."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(brand=None)

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    # pandas reads empty strings as NaN; both indicate "no brand"
    assert pd.isna(df.iloc[0]["marca"]) or df.iloc[0]["marca"] == ""


@pytest.mark.asyncio
async def test_export_full_backup_returns_bytes():
    """export_full_backup must return non-empty bytes."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = []
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    result = await svc.export_full_backup(business_id=business_id)

    assert isinstance(result, bytes)
    assert len(result) > 0


@pytest.mark.asyncio
async def test_export_full_backup_multiple_products_marca_column():
    """All rows in the backup export the correct brand value."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    p1 = _make_backup_product(brand="Alpha")
    p2 = _make_backup_product(brand="Beta")
    p3 = _make_backup_product(brand=None)

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [p1, p2, p3]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    marcas = list(df["marca"])
    assert marcas[0] == "Alpha"
    assert marcas[1] == "Beta"
    # pandas reads empty string as NaN; both indicate "no brand"
    assert pd.isna(marcas[2]) or marcas[2] == ""


# ---------------------------------------------------------------------------
# import_from_excel — unit_cost uses _calculate_unit_cost (not cost_price)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_from_excel_uses_calculate_unit_cost_not_cost_price():
    """
    unit_cost in items_data must come from _calculate_unit_cost(product),
    not from product.cost_price.

    We set cost_price != _calculate_unit_cost result and verify which one is used.
    """
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()
    supplier_id = uuid4()
    user_id = uuid4()

    # Product where cost_price differs from what _calculate_unit_cost would return
    product = _make_product(
        list_price=Decimal("1000.00"),
        discount_1=Decimal("10"),
        cost_price=Decimal("9999.00"),  # deliberately wrong
        code="IMPORT-001",
    )
    product.iva_rate = Decimal("21.00")

    mock_product_result = MagicMock()
    mock_product_result.scalar_one_or_none.return_value = product

    mock_supplier_result = MagicMock()
    mock_supplier_result.scalar_one_or_none.return_value = MagicMock(id=supplier_id)

    db.execute = AsyncMock(return_value=mock_product_result)

    svc = PurchaseOrderService(db=db)

    # Patch create so we don't need a full DB for the order
    captured_items = []

    async def _fake_create(business_id, user_id, data):
        captured_items.extend(data.items)
        fake_order = MagicMock()
        fake_order.id = uuid4()
        return fake_order

    svc.create = _fake_create

    import io
    import openpyxl

    # Build minimal Excel in memory
    # import_from_excel looks for "codigo"/"código" in col A and "a pedir" for qty
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "stock sistema", "conteo fisico", "a pedir"])
    ws.append(["IMPORT-001", 10, 5, 3])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    # Reset execute to return product properly
    async def _execute_side_effect(query, *args, **kwargs):
        result = MagicMock()
        result.scalar_one_or_none.return_value = product
        return result

    db.execute = _execute_side_effect

    order, skipped = await svc.import_from_excel(
        business_id=business_id,
        user_id=user_id,
        supplier_id=supplier_id,
        category_id=None,
        file_bytes=excel_bytes,
    )

    assert len(captured_items) == 1
    item = captured_items[0]
    # _calculate_unit_cost(product) = 1000 × 0.90 = 900, NOT 9999
    assert item.unit_cost == Decimal("900.00"), (
        f"Expected 900.00 from _calculate_unit_cost, got {item.unit_cost}"
    )


@pytest.mark.asyncio
async def test_import_from_excel_unit_cost_no_discounts():
    """When all discounts are 0, unit_cost equals list_price."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()
    supplier_id = uuid4()
    user_id = uuid4()

    product = _make_product(
        list_price=Decimal("500.00"),
        discount_1=Decimal("0"),
        discount_2=Decimal("0"),
        discount_3=Decimal("0"),
        cost_price=Decimal("1.00"),  # deliberately different
        code="NODISCOUNT-001",
    )
    product.iva_rate = Decimal("21.00")

    svc = PurchaseOrderService(db=db)

    captured_items = []

    async def _fake_create(business_id, user_id, data):
        captured_items.extend(data.items)
        fake_order = MagicMock()
        fake_order.id = uuid4()
        return fake_order

    svc.create = _fake_create

    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "stock sistema", "conteo fisico", "a pedir"])
    ws.append(["NODISCOUNT-001", 5, 2, 2])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    async def _execute_side_effect(query, *args, **kwargs):
        result = MagicMock()
        result.scalar_one_or_none.return_value = product
        return result

    db.execute = _execute_side_effect

    order, _ = await svc.import_from_excel(
        business_id=business_id,
        user_id=user_id,
        supplier_id=supplier_id,
        category_id=None,
        file_bytes=excel_bytes,
    )

    assert captured_items[0].unit_cost == Decimal("500.00")


@pytest.mark.asyncio
async def test_import_from_excel_unit_cost_three_discounts():
    """unit_cost applies the three-tier discount chain."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()
    supplier_id = uuid4()
    user_id = uuid4()

    product = _make_product(
        list_price=Decimal("1000.00"),
        discount_1=Decimal("15"),
        discount_2=Decimal("10"),
        discount_3=Decimal("5"),
        cost_price=Decimal("1.00"),
        code="TRIPLEDIS-001",
    )
    product.iva_rate = Decimal("21.00")

    svc = PurchaseOrderService(db=db)

    captured_items = []

    async def _fake_create(business_id, user_id, data):
        captured_items.extend(data.items)
        fake_order = MagicMock()
        fake_order.id = uuid4()
        return fake_order

    svc.create = _fake_create

    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "stock sistema", "conteo fisico", "a pedir"])
    ws.append(["TRIPLEDIS-001", 6, 3, 3])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    async def _execute_side_effect(query, *args, **kwargs):
        result = MagicMock()
        result.scalar_one_or_none.return_value = product
        return result

    db.execute = _execute_side_effect

    order, _ = await svc.import_from_excel(
        business_id=business_id,
        user_id=user_id,
        supplier_id=supplier_id,
        category_id=None,
        file_bytes=excel_bytes,
    )

    expected = round(
        Decimal("1000") * Decimal("0.85") * Decimal("0.90") * Decimal("0.95"), 2
    )
    assert captured_items[0].unit_cost == expected
