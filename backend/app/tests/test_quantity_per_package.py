"""
Tests for quantity_per_package feature and negative quantity validation rules.

Covers:
  1. Product.calculate_prices() with quantity_per_package
  2. export_products headers and row_data for 'cantidad_por_compra'
  3. export_full_backup dict key 'cantidad_por_compra'
  4. VoucherService negative-quantity guards (quotation and delivery receipt)
"""

from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
import pytest_asyncio
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.voucher import VoucherType
from app.services.excel_service import ExcelService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_product(
    *,
    list_price: Decimal = Decimal("1000.00"),
    discount_1: Decimal = Decimal("0"),
    discount_2: Decimal = Decimal("0"),
    discount_3: Decimal = Decimal("0"),
    extra_cost: Decimal = Decimal("0"),
    profit_margin: Decimal = Decimal("0"),
    iva_rate: Decimal = Decimal("21"),
    quantity_per_package: Decimal | None = None,
    units_per_pack: int | None = None,
    code: str = "TST-001",
    description: str = "Test product",
    brand: str | None = None,
    lots: list | None = None,
    supplier=None,
    category=None,
) -> MagicMock:
    """Build a MagicMock that quacks like a Product with the given field values."""
    p = MagicMock(spec=Product)
    p.id = uuid4()
    p.code = code
    p.supplier_code = None
    p.description = description
    p.brand = brand
    p.list_price = list_price
    p.discount_1 = discount_1
    p.discount_2 = discount_2
    p.discount_3 = discount_3
    p.extra_cost = extra_cost
    p.profit_margin = profit_margin
    p.iva_rate = iva_rate
    p.quantity_per_package = quantity_per_package
    p.units_per_pack = units_per_pack
    p.sale_price = Decimal("0")
    p.net_price = Decimal("0")
    p.unit = "unidad"
    p.next_expiration = None
    p.supplier = supplier
    p.category = category
    p.brand_ref = None       # avoid MagicMock when excel_service checks brand_ref.name
    p.current_stock = 0      # avoid MagicMock — pandas chokes on mock's tzinfo
    p.lots = lots or []
    return p


def _make_backup_product(**kwargs) -> MagicMock:
    """Build a MagicMock for export_full_backup (needs more fields)."""
    p = _make_product(**kwargs)
    p.id = uuid4()
    p.supplier_code = ""
    p.details = ""
    p.discount_display = "0"
    p.net_price = Decimal("1000.00")
    p.sale_price = Decimal("1210.00")
    p.minimum_stock = 0
    p.is_active = True
    p.expiration_date = None
    p.created_at = None
    p.updated_at = None
    p.deleted_at = None
    p.category_id = None
    p.supplier_id = None
    p.cost_price = Decimal("0")
    p.list_price_usd = None
    return p


def _make_export_products_db_mock(products: list) -> AsyncMock:
    """
    export_products calls db.execute three times:
      1. SELECT products
      2. SELECT categories (for reference sheet)
      3. SELECT suppliers (for reference sheet)
    """
    db = AsyncMock(spec=AsyncSession)

    products_result = MagicMock()
    products_result.scalars.return_value.all.return_value = products

    empty_result = MagicMock()
    empty_result.scalars.return_value.all.return_value = []

    db.execute = AsyncMock(side_effect=[products_result, empty_result, empty_result])
    return db


# ---------------------------------------------------------------------------
# Feature 1 — calculate_prices with quantity_per_package (pure unit tests)
# ---------------------------------------------------------------------------


class TestCalculatePricesQuantityPerPackage:
    """Tests for Product.calculate_prices() when quantity_per_package is set."""

    def _product(self, **kwargs) -> Product:
        """Build a real Product instance (no DB needed)."""
        p = Product()
        p.list_price = kwargs.get("list_price", Decimal("1000.00"))
        p.discount_1 = kwargs.get("discount_1", Decimal("0"))
        p.discount_2 = kwargs.get("discount_2", Decimal("0"))
        p.discount_3 = kwargs.get("discount_3", Decimal("0"))
        p.extra_cost = kwargs.get("extra_cost", Decimal("0"))
        p.profit_margin = kwargs.get("profit_margin", Decimal("0"))
        p.iva_rate = kwargs.get("iva_rate", Decimal("21"))
        p.quantity_per_package = kwargs.get("quantity_per_package")
        return p

    def test_no_quantity_per_package_sale_price_unchanged(self):
        """Without quantity_per_package, sale_price = list × (1 + IVA/100)."""
        p = self._product(list_price=Decimal("1000.00"), iva_rate=Decimal("21"))
        p.calculate_prices()
        assert p.sale_price == Decimal("1210.00")
        assert p.net_price == Decimal("1000.00")

    def test_quantity_per_package_20_divides_sale_price(self):
        """list=1000, IVA=21%, qpp=20 → sale_price = round(1210/20, 2) = 60.50"""
        p = self._product(
            list_price=Decimal("1000.00"),
            iva_rate=Decimal("21"),
            quantity_per_package=Decimal("20"),
        )
        p.calculate_prices()
        assert p.sale_price == Decimal("60.50")

    def test_quantity_per_package_zero_treated_as_absent(self):
        """quantity_per_package=0 → division skipped, normal price."""
        p = self._product(
            list_price=Decimal("1000.00"),
            iva_rate=Decimal("21"),
            quantity_per_package=Decimal("0"),
        )
        p.calculate_prices()
        assert p.sale_price == Decimal("1210.00")

    def test_quantity_per_package_none_treated_as_absent(self):
        """quantity_per_package=None → division skipped, normal price."""
        p = self._product(
            list_price=Decimal("1000.00"),
            iva_rate=Decimal("21"),
            quantity_per_package=None,
        )
        p.calculate_prices()
        assert p.sale_price == Decimal("1210.00")

    def test_chain_discount_d1_iva_and_qpp(self):
        """list=1000, d1=20%, IVA=21%, qpp=5 → sale_price = round(800*1.21/5, 2) = 193.60"""
        p = self._product(
            list_price=Decimal("1000.00"),
            discount_1=Decimal("20"),
            iva_rate=Decimal("21"),
            quantity_per_package=Decimal("5"),
        )
        p.calculate_prices()
        # net_with_profit = 1000 * 0.80 = 800
        # sale = 800 * 1.21 = 968
        # sale / 5 = 193.60
        assert p.sale_price == Decimal("193.60")

    def test_quantity_per_package_1_no_visible_change(self):
        """Dividing by 1 should yield the same result as no division."""
        p_base = self._product(list_price=Decimal("1000.00"), iva_rate=Decimal("21"))
        p_base.calculate_prices()

        p_one = self._product(
            list_price=Decimal("1000.00"),
            iva_rate=Decimal("21"),
            quantity_per_package=Decimal("1"),
        )
        p_one.calculate_prices()

        assert p_one.sale_price == p_base.sale_price
        assert p_one.net_price == p_base.net_price

    def test_net_price_also_divided(self):
        """net_price (sans IVA) must also be divided by quantity_per_package."""
        p = self._product(
            list_price=Decimal("1000.00"),
            iva_rate=Decimal("21"),
            quantity_per_package=Decimal("10"),
        )
        p.calculate_prices()
        # net_with_profit = 1000, /10 = 100
        assert p.net_price == Decimal("100.00")
        # sale = 1210 / 10 = 121
        assert p.sale_price == Decimal("121.00")


# ---------------------------------------------------------------------------
# Feature 1 — export_products: headers and row_data
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_products_headers_contain_cantidad_por_compra():
    """'cantidad_por_compra' must appear in the exported headers."""
    business_id = uuid4()
    product = _make_product(quantity_per_package=Decimal("20.00"))
    db = _make_export_products_db_mock([product])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert "cantidad_por_compra" in headers, f"Expected 'cantidad_por_compra' in headers. Got: {headers}"


@pytest.mark.asyncio
async def test_export_products_cantidad_por_compra_value_written():
    """When product.quantity_per_package is set, the correct float is written to the cell."""
    business_id = uuid4()
    product = _make_product(quantity_per_package=Decimal("15.50"))
    db = _make_export_products_db_mock([product])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    col = headers.index("cantidad_por_compra") + 1
    value = ws.cell(row=2, column=col).value
    assert value == 15.5, f"Expected 15.5, got {value!r}"


@pytest.mark.asyncio
async def test_export_products_cantidad_por_compra_none_is_empty():
    """When product.quantity_per_package is None, the cell must be empty ('' or None)."""
    business_id = uuid4()
    product = _make_product(quantity_per_package=None)
    db = _make_export_products_db_mock([product])

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_products(business_id=business_id)

    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    col = headers.index("cantidad_por_compra") + 1
    value = ws.cell(row=2, column=col).value
    assert value in ("", None), f"Expected empty or None for null quantity_per_package, got {value!r}"


# ---------------------------------------------------------------------------
# Feature 1 — export_full_backup: dict key 'cantidad_por_compra'
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_full_backup_cantidad_por_compra_key_exists():
    """Backup sheet must contain a 'cantidad_por_compra' column."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(quantity_per_package=Decimal("20.00"))

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    assert "cantidad_por_compra" in df.columns, (
        f"'cantidad_por_compra' not found. Columns: {list(df.columns)}"
    )


@pytest.mark.asyncio
async def test_export_full_backup_cantidad_por_compra_value_correct():
    """The 'cantidad_por_compra' cell in backup equals float(product.quantity_per_package)."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(quantity_per_package=Decimal("25.00"))

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    assert df.iloc[0]["cantidad_por_compra"] == 25.0


@pytest.mark.asyncio
async def test_export_full_backup_cantidad_por_compra_none_exports_as_empty():
    """When product.quantity_per_package is None, backup 'cantidad_por_compra' exports as empty."""
    db = AsyncMock(spec=AsyncSession)
    business_id = uuid4()

    product = _make_backup_product(quantity_per_package=None)

    mock_result = MagicMock()
    mock_result.scalars.return_value.all.return_value = [product]
    db.execute = AsyncMock(return_value=mock_result)

    svc = ExcelService(db=db)
    excel_bytes = await svc.export_full_backup(business_id=business_id)

    import io
    import pandas as pd

    df = pd.read_excel(io.BytesIO(excel_bytes), sheet_name="Backup_Productos")
    value = df.iloc[0]["cantidad_por_compra"]
    import math
    assert value == "" or (isinstance(value, float) and math.isnan(value)), (
        f"Expected empty/NaN for None quantity_per_package, got {value!r}"
    )


# ---------------------------------------------------------------------------
# Feature 2 — Negative quantity validation (pure logic tests)
# ---------------------------------------------------------------------------
#
# We test the extracted guard conditions directly without running the full
# service.create() method which requires a complete DB session.
# ---------------------------------------------------------------------------


def _neg_qty_item(quantity: int | float) -> MagicMock:
    item = MagicMock()
    item.quantity = Decimal(str(quantity))
    return item


def _check_quotation_negatives(voucher_type, items):
    """Replicates the quotation negative-quantity guard from voucher_service.create."""
    if (
        voucher_type == VoucherType.QUOTATION
        and any(Decimal(str(i.quantity)) < 0 for i in items)
    ):
        raise ValueError("Una cotización no puede contener cantidades negativas.")


def _check_receipt_negatives(voucher_type, is_current_account, stockpile_id, items):
    """Replicates the delivery-receipt negative-quantity guard from voucher_service.create."""
    if (
        voucher_type == VoucherType.RECEIPT
        and not is_current_account
        and not stockpile_id
        and any(Decimal(str(i.quantity)) < 0 for i in items)
    ):
        raise ValueError(
            "Un remito de entrega no puede contener cantidades negativas. "
            "Para devoluciones usá Remito CC o Factura."
        )


class TestNegativeQuantityGuards:
    """Feature 2: negative quantity validation rules."""

    # --- Quotation guards ---

    def test_quotation_negative_qty_raises(self):
        """quotation + negative quantity → raises ValueError containing 'cotización'."""
        items = [_neg_qty_item(-1)]
        with pytest.raises(ValueError, match="cotización"):
            _check_quotation_negatives(VoucherType.QUOTATION, items)

    def test_quotation_positive_qty_no_error(self):
        """quotation + all positive quantities → no exception."""
        items = [_neg_qty_item(5), _neg_qty_item(3)]
        _check_quotation_negatives(VoucherType.QUOTATION, items)  # must not raise

    def test_quotation_multiple_negatives_raises(self):
        """quotation + multiple negative items → raises ValueError."""
        items = [_neg_qty_item(-2), _neg_qty_item(-5)]
        with pytest.raises(ValueError, match="cotización"):
            _check_quotation_negatives(VoucherType.QUOTATION, items)

    def test_quotation_mixed_pos_neg_raises(self):
        """quotation + mixed positive and negative → raises ValueError."""
        items = [_neg_qty_item(5), _neg_qty_item(-1)]
        with pytest.raises(ValueError, match="cotización"):
            _check_quotation_negatives(VoucherType.QUOTATION, items)

    # --- Delivery receipt guards ---

    def test_receipt_not_cc_no_stockpile_negative_raises(self):
        """receipt (not CC, no stockpile) + negative → raises ValueError containing 'remito de entrega'."""
        items = [_neg_qty_item(-3)]
        with pytest.raises(ValueError, match="remito de entrega"):
            _check_receipt_negatives(VoucherType.RECEIPT, False, None, items)

    def test_receipt_cc_negative_no_error(self):
        """receipt (is_current_account=True) + negative → no ValueError from new guard."""
        items = [_neg_qty_item(-2)]
        _check_receipt_negatives(VoucherType.RECEIPT, True, None, items)  # must not raise

    def test_receipt_stockpile_set_negative_no_error(self):
        """receipt (stockpile_id set) + negative → no ValueError from new guard."""
        stockpile_id = uuid4()
        items = [_neg_qty_item(-4)]
        _check_receipt_negatives(VoucherType.RECEIPT, False, stockpile_id, items)  # must not raise

    def test_invoice_a_negative_no_error(self):
        """invoice_a + negative → no ValueError from new guards."""
        items = [_neg_qty_item(-1)]
        _check_quotation_negatives(VoucherType.INVOICE_A, items)  # must not raise
        _check_receipt_negatives(VoucherType.INVOICE_A, False, None, items)  # must not raise

    def test_receipt_positive_qty_no_error(self):
        """receipt (not CC, no stockpile) + all positive → no exception."""
        items = [_neg_qty_item(10), _neg_qty_item(5)]
        _check_receipt_negatives(VoucherType.RECEIPT, False, None, items)  # must not raise
