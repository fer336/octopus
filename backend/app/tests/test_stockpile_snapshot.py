"""
Tests para el feature de Snapshots de Precios de Acopio (PR #1).

Cubre:
- Creación de snapshots al hacer create_by_amount()
- Contenido y estructura de snapshots
- Generación de Excel con snapshots
- Endpoint de descarga Excel
- Resolución de precios desde snapshots en retiros (voucher_service)
- Legacy fallback para acopios sin snapshots
"""
import asyncio
import io
from datetime import date
from decimal import Decimal
from uuid import UUID

import httpx
import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.business import Business
from app.models.client import Client
from app.models.product import Product
from app.models.stockpile import StockpilePriceSnapshot
from app.models.tenant_membership import TenantMembership
from app.models.user import User
from app.models.voucher import VoucherType
from app.schemas.voucher import VoucherCreate, VoucherItemCreate
from app.tests.conftest import make_auth_header


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _create_products(db: AsyncSession, business_id: UUID) -> list[Product]:
    """Crea productos de prueba activos."""
    products_data = [
        # (code, description, list_price, iva_rate, is_active)
        ("SNAP-A", "Producto A", Decimal("100.00"), Decimal("21.00"), True),
        ("SNAP-B", "Producto B", Decimal("80.00"), Decimal("10.50"), True),
        ("SNAP-INACTIVE", "Producto Inactivo", Decimal("20.00"), Decimal("21.00"), False),
    ]
    from app.models.product_lot import ProductLot

    products = []
    for code, desc, list_price, iva, active in products_data:
        p = Product(
            description=desc,
            code=code,
            list_price=list_price,
            cost_price=list_price * Decimal("0.6"),
            iva_rate=iva,
            business_id=business_id,
            is_active=active,
        )
        p.calculate_prices()
        db.add(p)
        products.append(p)

    await db.flush()

    # Crear lotes para productos activos (necesario para stock > 0)
    for p in products:
        if p.is_active:
            lot = ProductLot(
                product_id=p.id,
                business_id=business_id,
                quantity=100,
                initial_quantity=100,
                cost_price=p.cost_price or Decimal("0"),
            )
            db.add(lot)

    await db.commit()
    for p in products:
        await db.refresh(p)
    return products


async def _create_client(db: AsyncSession, business_id: UUID) -> Client:
    """Crea un cliente de prueba con su tipo de cliente."""
    from app.models.client_type import ClientType

    # Crear tipo de cliente primero
    ct = ClientType(
        business_id=business_id,
        name="Test Type",
    )
    db.add(ct)
    await db.flush()

    c = Client(
        business_id=business_id,
        name="Cliente Test",
        email="cliente@test.com",
        client_type_id=ct.id,
        document_type="DNI",
        document_number="12345678",
        tax_condition="Consumidor Final",
    )
    db.add(c)
    await db.commit()
    await db.refresh(c)
    return c


# ---------------------------------------------------------------------------
# Tests: Snapshot creation via create_by_amount
# ---------------------------------------------------------------------------


class TestSnapshotCreation:
    """Snapshots se crean automáticamente al hacer create_by_amount()."""

    @pytest.mark.asyncio
    async def test_snapshots_created_on_create_by_amount(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Verifica que al crear acopio por monto se generan snapshots para
        productos activos."""
        from app.services.stockpile_service import StockpileService

        products = await _create_products(db, business_a.id)
        client = await _create_client(db, business_a.id)
        active_products = [p for p in products if p.is_active]
        inactive_product = [p for p in products if not p.is_active][0]

        service = StockpileService(db)
        stockpile, item = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client.id,
            created_by=user_a.id,
            name="Acopio Test Snapshots",
            description="Test para snapshots",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("10000.00"),
            discount_percent=Decimal("0"),
            billing_client_id=None,
        )

        # Verificar snapshots creados
        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
                StockpilePriceSnapshot.deleted_at.is_(None),
            )
        )
        snapshots = list(result.scalars().all())

        assert len(snapshots) == len(active_products), (
            f"Expected {len(active_products)} snapshots, got {len(snapshots)}"
        )

        snapshot_product_ids = {s.product_id for s in snapshots}
        for ap in active_products:
            assert ap.id in snapshot_product_ids, (
                f"Active product {ap.code} missing from snapshots"
            )

        # Producto inactivo NO debe estar en snapshots
        assert inactive_product.id not in snapshot_product_ids, (
            "Inactive product should not appear in snapshots"
        )

    @pytest.mark.asyncio
    async def test_snapshot_content(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Verifica que los valores en el snapshot sean correctos."""
        from app.services.stockpile_service import StockpileService

        products = await _create_products(db, business_a.id)
        client = await _create_client(db, business_a.id)

        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client.id,
            created_by=user_a.id,
            name="Acopio Verificación",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
                StockpilePriceSnapshot.product_id == products[0].id,
            )
        )
        snapshot = result.scalar_one_or_none()
        assert snapshot is not None, "Snapshot for producto A should exist"

        product = products[0]
        expected_iva = product.net_price * product.iva_rate / Decimal("100")
        expected_iva = expected_iva.quantize(Decimal("0.01"))
        expected_total = (product.net_price + expected_iva).quantize(Decimal("0.01"))

        assert snapshot.code == product.code
        assert snapshot.description == product.description
        assert snapshot.price_without_iva == product.net_price
        assert snapshot.iva_rate == product.iva_rate
        assert snapshot.iva_amount == expected_iva
        assert snapshot.price_with_iva == expected_total
        assert snapshot.frozen_at is not None

    @pytest.mark.asyncio
    async def test_no_snapshots_for_by_product_stockpile(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Acopios por producto NO generan snapshots."""
        from app.services.stockpile_service import StockpileService

        products = await _create_products(db, business_a.id)
        client = await _create_client(db, business_a.id)
        active = [p for p in products if p.is_active]

        service = StockpileService(db)
        items_data = [{"product_id": p.id, "quantity": Decimal("1")} for p in active[:1]]

        stockpile, _ = await service.create(
            business_id=business_a.id,
            client_id=client.id,
            created_by=user_a.id,
            name="Acopio por Producto",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            billing_client_id=None,
            items_data=items_data,
        )

        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
            )
        )
        snapshots = list(result.scalars().all())
        assert len(snapshots) == 0, (
            "By-product stockpiles should NOT create snapshots"
        )

    @pytest.mark.asyncio
    async def test_snapshots_created_via_endpoint(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        membership_a: TenantMembership,
    ):
        """El endpoint create_by_amount ejecuta sin error y persiste snapshots."""
        products = await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        headers = make_auth_header(user_a)

        payload = {
            "client_id": str(client_entity.id),
            "name": "Acopio Endpoint Snapshots",
            "currency": "ARS",
            "exchange_rate": 1.0,
            "amount": 10000.0,
            "discount_percent": 0,
        }
        response = await client.post(
            "/api/tenant/stockpiles/by-amount",
            json=payload,
            headers=headers,
        )
        assert response.status_code == 201, f"Status {response.status_code}: {response.text[:500]}"
        data = response.json()

        # Verificar snapshots en DB
        stockpile_id = UUID(data["id"])
        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile_id,
                StockpilePriceSnapshot.deleted_at.is_(None),
            )
        )
        snapshots = list(result.scalars().all())
        active_count = len([p for p in products if p.is_active])
        assert len(snapshots) == active_count


# ---------------------------------------------------------------------------
# Tests: Excel generation
# ---------------------------------------------------------------------------


class TestExcelGeneration:
    """Generación de Excel a partir de snapshots."""

    @pytest.mark.asyncio
    async def test_generate_excel_contains_snapshots(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Verifica que el Excel generado contenga todos los snapshots."""
        from app.services.stockpile_service import StockpileService
        from app.services.stockpile_snapshot_service import StockpileSnapshotService

        products = await _create_products(db, business_a.id)
        client = await _create_client(db, business_a.id)
        active = [p for p in products if p.is_active]

        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client.id,
            created_by=user_a.id,
            name="Acopio Excel Test",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        snapshot_service = StockpileSnapshotService(db)
        snapshots = await snapshot_service.get_snapshots(stockpile.id)
        assert len(snapshots) == len(active)

        excel_bytes = snapshot_service.generate_excel(snapshots, stockpile.name)
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb.active

        # Verificar título
        assert ws["A1"].value is not None
        assert stockpile.name in ws["A1"].value

        # Verificar headers (fila 3)
        headers = [ws.cell(row=3, column=c).value for c in range(1, 8)]
        assert "Código" in headers
        assert "Descripción" in headers
        assert "Precio sin IVA" in headers
        assert "IVA %" in headers
        assert "IVA $" in headers
        assert "Precio final con IVA" in headers
        assert "Fecha de congelamiento" in headers

        # Verificar datos (desde fila 4)
        for i, snap in enumerate(snapshots):
            row = 4 + i
            assert ws.cell(row=row, column=1).value == snap.code
            assert ws.cell(row=row, column=2).value == snap.description
            assert float(ws.cell(row=row, column=3).value) == float(snap.price_without_iva)
            assert float(ws.cell(row=row, column=4).value) == float(snap.iva_rate)
            assert float(ws.cell(row=row, column=6).value) == float(snap.price_with_iva)

        # Verificar autofilter
        assert ws.auto_filter.ref is not None

    @pytest.mark.asyncio
    async def test_generate_excel_no_snapshots(
        self,
        db: AsyncSession,
        business_a: Business,
    ):
        """Excel sin snapshots debe funcionar."""
        from app.services.stockpile_snapshot_service import StockpileSnapshotService

        snapshot_service = StockpileSnapshotService(db)
        excel_bytes = snapshot_service.generate_excel([], "Acopio Vacío")
        wb = openpyxl.load_workbook(excel_bytes)
        ws = wb.active
        assert ws is not None


# ---------------------------------------------------------------------------
# Tests: Excel download endpoint
# ---------------------------------------------------------------------------


class TestExcelEndpoint:
    """Endpoint GET /stockpiles/{id}/price-snapshot/excel."""

    @pytest.mark.asyncio
    async def test_download_excel_success(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        membership_a: TenantMembership,
    ):
        """Descarga exitosa de Excel para acopio por monto."""
        from app.services.stockpile_service import StockpileService

        await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        headers = make_auth_header(user_a)

        # Crear acopio por monto
        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client_entity.id,
            created_by=user_a.id,
            name="Acopio Download",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        response = await client.get(
            f"/api/tenant/stockpiles/{stockpile.id}/price-snapshot/excel",
            headers=headers,
        )
        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response.headers.get("content-disposition", "")

        # Validar que sea un Excel válido
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws is not None

    @pytest.mark.asyncio
    async def test_download_excel_by_product_stockpile_404(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        membership_a: TenantMembership,
    ):
        """Acopio por producto NO tiene snapshots → 404."""
        from app.services.stockpile_service import StockpileService

        products = await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        headers = make_auth_header(user_a)
        active = [p for p in products if p.is_active]

        service = StockpileService(db)
        items_data = [{"product_id": p.id, "quantity": Decimal("1")} for p in active[:1]]

        stockpile, _ = await service.create(
            business_id=business_a.id,
            client_id=client_entity.id,
            created_by=user_a.id,
            name="Acopio Prod",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            billing_client_id=None,
            items_data=items_data,
        )

        response = await client.get(
            f"/api/tenant/stockpiles/{stockpile.id}/price-snapshot/excel",
            headers=headers,
        )
        assert response.status_code == 404

    @pytest.mark.asyncio
    async def test_download_excel_not_found(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        membership_a: TenantMembership,
    ):
        """Stockpile inexistente → 404."""
        headers = make_auth_header(user_a)
        fake_id = "00000000-0000-0000-0000-000000000000"
        response = await client.get(
            f"/api/tenant/stockpiles/{fake_id}/price-snapshot/excel",
            headers=headers,
        )
        assert response.status_code == 404

    @pytest.mark.asyncio
    async def test_download_excel_multitenant_isolation(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_b: Business,
        user_b: User,
        membership_b: TenantMembership,
    ):
        """Stockpile de tenant B no visible por tenant A → 404."""
        from app.services.stockpile_service import StockpileService
        from app.models.client_type import ClientType

        # Crear cliente y productos para tenant B
        ct_b = ClientType(business_id=business_b.id, name="Type B")
        db.add(ct_b)
        await db.flush()

        client_b = Client(
            name="Cliente B",
            business_id=business_b.id,
            client_type_id=ct_b.id,
            document_type="DNI",
            document_number="87654321",
            tax_condition="Consumidor Final",
        )
        db.add(client_b)
        await db.commit()
        await db.refresh(client_b)

        product_b = Product(
            description="Producto B",
            code="B-001",
            list_price=Decimal("100.00"),
            cost_price=Decimal("60.00"),
            iva_rate=Decimal("21.00"),
            business_id=business_b.id,
            is_active=True,
        )
        product_b.calculate_prices()
        db.add(product_b)
        await db.commit()

        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_b.id,
            client_id=client_b.id,
            created_by=user_b.id,
            name="Acopio B",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        # user_a (quien NO tiene acceso a tenant B) intenta acceder
        # y debe obtener 404 (el negocio existe pero no le pertenece)
        headers = make_auth_header(user_a)
        response = await client.get(
            f"/api/tenant/stockpiles/{stockpile.id}/price-snapshot/excel",
            headers=headers,
        )
        # user_a no tiene negocio asociado, obtiene 403
        assert response.status_code in (403, 404), (
            f"Expected 403 or 404, got {response.status_code}: {response.text[:200]}"
        )

    @pytest.mark.asyncio
    async def test_download_excel_accepts_internal_n8n_api_key(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        monkeypatch,
    ):
        """n8n puede descargar el Excel sin sesión usando la API key interna."""
        from app.services.stockpile_service import StockpileService

        monkeypatch.setenv("N8N_STOCKPILE_SNAPSHOT_API_KEY", "test-stockpile-key")
        await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)

        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client_entity.id,
            created_by=user_a.id,
            name="Acopio API Key",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        response = await client.get(
            f"/api/tenant/stockpiles/{stockpile.id}/price-snapshot/excel",
            headers={"X-N8N-Stockpile-API-Key": "test-stockpile-key"},
        )
        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ---------------------------------------------------------------------------
# Tests: n8n webhook dispatch
# ---------------------------------------------------------------------------


class TestStockpileWebhookDispatch:
    """Webhook fire-and-forget para envío de resumen de acopio por n8n."""

    @pytest.mark.asyncio
    async def test_dispatch_webhook_posts_expected_payload(
        self,
        db: AsyncSession,
        monkeypatch,
    ):
        """El servicio POSTea el contrato esperado hacia n8n."""
        from app.services.stockpile_snapshot_service import StockpileSnapshotService

        calls = []

        class MockAsyncClient:
            def __init__(self, timeout):
                self.timeout = timeout

            async def __aenter__(self):
                return self

            async def __aexit__(self, exc_type, exc, tb):
                return None

            async def post(self, url, json):
                calls.append({"url": url, "json": json, "timeout": self.timeout})
                return httpx.Response(
                    202,
                    request=httpx.Request("POST", url),
                )

        monkeypatch.setenv("N8N_STOCKPILE_WEBHOOK_URL", "https://n8n.test/webhook/acopio")
        monkeypatch.setenv("N8N_STOCKPILE_SNAPSHOT_API_KEY", "snapshot-secret")
        monkeypatch.setattr(
            "app.services.stockpile_snapshot_service.httpx.AsyncClient",
            MockAsyncClient,
        )

        stockpile_id = UUID("00000000-0000-0000-0000-000000000123")
        service = StockpileSnapshotService(db)
        result = await service.dispatch_webhook(
            stockpile_id=stockpile_id,
            stockpile_name="Obra Centro",
            stockpile_number="AC-0001",
            client_email="cliente@example.com",
            client_name="Juan Perez",
            business_name="Mi Ferretería",
            base_url="https://api.example.com/",
        )

        assert result == {"sent": True, "status_code": 202}
        assert len(calls) == 1
        assert calls[0]["url"] == "https://n8n.test/webhook/acopio"
        assert calls[0]["timeout"] == 5.0
        assert calls[0]["json"] == {
            "event": "stockpile_created",
            "stockpile_id": str(stockpile_id),
            "stockpile_name": "Obra Centro",
            "stockpile_number": "AC-0001",
            "client_email": "cliente@example.com",
            "client_name": "Juan Perez",
            "business_name": "Mi Ferretería",
            "snapshot_url": (
                f"https://api.example.com/api/tenant/stockpiles/{stockpile_id}"
                "/price-snapshot/excel"
            ),
            "auth_token": "snapshot-secret",
        }

    @pytest.mark.asyncio
    async def test_dispatch_webhook_timeout_does_not_crash(
        self,
        db: AsyncSession,
        monkeypatch,
    ):
        """Un timeout de n8n no debe romper la creación del acopio."""
        from app.services.stockpile_snapshot_service import StockpileSnapshotService

        class TimeoutAsyncClient:
            def __init__(self, timeout):
                self.timeout = timeout

            async def __aenter__(self):
                return self

            async def __aexit__(self, exc_type, exc, tb):
                return None

            async def post(self, url, json):
                raise httpx.TimeoutException("timeout")

        monkeypatch.setenv("N8N_STOCKPILE_WEBHOOK_URL", "https://n8n.test/webhook/acopio")
        monkeypatch.setattr(
            "app.services.stockpile_snapshot_service.httpx.AsyncClient",
            TimeoutAsyncClient,
        )

        service = StockpileSnapshotService(db)
        result = await service.dispatch_webhook(
            stockpile_id=UUID("00000000-0000-0000-0000-000000000124"),
            stockpile_name="Obra Timeout",
            stockpile_number="AC-0002",
            client_email="cliente@example.com",
            client_name="Juan Perez",
            business_name="Mi Ferretería",
            base_url="https://api.example.com",
        )

        assert result == {"sent": False, "reason": "timeout"}

    @pytest.mark.asyncio
    async def test_create_by_amount_schedules_webhook_after_success(
        self,
        client: AsyncClient,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
        membership_a: TenantMembership,
        monkeypatch,
    ):
        """El endpoint crea el acopio y dispara POST fire-and-forget a n8n."""
        calls = []

        class MockAsyncClient:
            def __init__(self, timeout):
                self.timeout = timeout

            async def __aenter__(self):
                return self

            async def __aexit__(self, exc_type, exc, tb):
                return None

            async def post(self, url, json):
                calls.append({"url": url, "json": json})
                return httpx.Response(200, request=httpx.Request("POST", url))

        monkeypatch.setenv("N8N_STOCKPILE_WEBHOOK_URL", "https://n8n.test/webhook/acopio")
        monkeypatch.setenv("N8N_STOCKPILE_SNAPSHOT_API_KEY", "snapshot-secret")
        monkeypatch.setattr(
            "app.services.stockpile_snapshot_service.httpx.AsyncClient",
            MockAsyncClient,
        )

        await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        headers = make_auth_header(user_a)

        response = await client.post(
            "/api/tenant/stockpiles/by-amount",
            json={
                "client_id": str(client_entity.id),
                "name": "Acopio Webhook Endpoint",
                "currency": "ARS",
                "exchange_rate": 1.0,
                "amount": 10000.0,
                "discount_percent": 0,
            },
            headers=headers,
        )
        assert response.status_code == 201, response.text[:500]

        for _ in range(10):
            if calls:
                break
            await asyncio.sleep(0.01)

        assert len(calls) == 1
        payload = calls[0]["json"]
        assert payload["event"] == "stockpile_created"
        assert payload["stockpile_id"] == response.json()["id"]
        assert payload["stockpile_name"] == "Acopio Webhook Endpoint"
        assert payload["stockpile_number"].startswith("AC-")
        assert payload["client_email"] == "cliente@test.com"
        assert payload["client_name"] == "Cliente Test"
        assert payload["business_name"] == "Tenant A"
        assert payload["auth_token"] == "snapshot-secret"
        assert payload["snapshot_url"].endswith(
            f"/api/tenant/stockpiles/{response.json()['id']}/price-snapshot/excel"
        )


# ---------------------------------------------------------------------------
# Tests: Withdrawal price resolution from snapshots
# ---------------------------------------------------------------------------


class TestWithdrawalPriceResolution:
    """Resolución de precios desde snapshots al retirar (voucher_service)."""

    @pytest.mark.asyncio
    async def test_snapshot_price_used_in_voucher(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Al crear un RECEIPT con stockpile_id, los precios se resuelven
        desde snapshots en lugar del catálogo."""
        from app.services.stockpile_service import StockpileService
        from app.services.voucher_service import VoucherService

        products = await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        active = [p for p in products if p.is_active]

        # Crear acopio por monto → se generan snapshots
        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client_entity.id,
            created_by=user_a.id,
            name="Acopio Retiro Snapshots",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("1000.00"),
            discount_percent=Decimal("0"),
        )

        # Crear un voucher (RECEIPT) referenciando el stockpile
        # El unit_price enviado por frontend debe ser sobreescrito por el snapshot
        voucher_service = VoucherService(db)
        voucher = await voucher_service.create(
            business_id=business_a.id,
            data=VoucherCreate(
                client_id=client_entity.id,
                voucher_type=VoucherType.RECEIPT,
                date=date(2026, 5, 24),
                items=[
                    VoucherItemCreate(
                        product_id=active[0].id,
                        quantity=Decimal("1"),
                        unit_price=Decimal("999.99"),  # Frontend price, será sobreescrito
                    ),
                ],
                stockpile_id=stockpile.id,
                show_prices=True,
            ),
            user_id=user_a.id,
        )

        await db.refresh(voucher, ["items"])
        assert len(voucher.items) == 1
        item = voucher.items[0]

        # El precio debe ser el del snapshot, no el catálogo ni el enviado
        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
                StockpilePriceSnapshot.product_id == active[0].id,
            )
        )
        snapshot = result.scalar_one()
        assert float(item.unit_price) == float(snapshot.price_without_iva), (
            f"Expected snapshot price {snapshot.price_without_iva}, "
            f"got {item.unit_price} (catalog: {active[0].net_price})"
        )
        assert float(item.unit_price) != 999.99, (
            "Price should have been overridden by snapshot, not frontend value"
        )

    @pytest.mark.asyncio
    async def test_voucher_without_stockpile_uses_catalog_price(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Voucher sin stockpile_id no debe resolver desde snapshots."""
        from app.services.voucher_service import VoucherService

        products = await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        active = [p for p in products if p.is_active]

        voucher_service = VoucherService(db)
        voucher = await voucher_service.create(
            business_id=business_a.id,
            data=VoucherCreate(
                client_id=client_entity.id,
                voucher_type=VoucherType.RECEIPT,
                date=date(2026, 5, 24),
                items=[
                    VoucherItemCreate(
                        product_id=active[0].id,
                        quantity=Decimal("1"),
                        unit_price=Decimal("150.00"),
                    ),
                ],
                stockpile_id=None,
                show_prices=True,
            ),
            user_id=user_a.id,
        )

        await db.refresh(voucher, ["items"])
        item = voucher.items[0]
        assert float(item.unit_price) == 150.0

    @pytest.mark.asyncio
    async def test_by_product_stockpile_uses_catalog_price(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Acopio por producto usa precio del catálogo (no hay snapshots)."""
        from app.services.stockpile_service import StockpileService
        from app.services.voucher_service import VoucherService

        products = await _create_products(db, business_a.id)
        client_entity = await _create_client(db, business_a.id)
        active = [p for p in products if p.is_active]

        # Acopio por producto (description=None → by-product)
        service = StockpileService(db)
        items_data = [{"product_id": active[0].id, "quantity": Decimal("2")}]
        stockpile, _ = await service.create(
            business_id=business_a.id,
            client_id=client_entity.id,
            created_by=user_a.id,
            name="Acopio Producto",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            billing_client_id=None,
            items_data=items_data,
        )

        # Precio enviado por frontend (debería respetarse al no haber snapshots)
        sent_price = Decimal("155.00")

        voucher_service = VoucherService(db)
        voucher = await voucher_service.create(
            business_id=business_a.id,
            data=VoucherCreate(
                client_id=client_entity.id,
                voucher_type=VoucherType.RECEIPT,
                date=date(2026, 5, 24),
                items=[
                    VoucherItemCreate(
                        product_id=active[0].id,
                        quantity=Decimal("1"),
                        unit_price=sent_price,
                    ),
                ],
                stockpile_id=stockpile.id,
                show_prices=True,
            ),
            user_id=user_a.id,
        )

        await db.refresh(voucher, ["items"])
        item = voucher.items[0]
        # Para by-product, el precio NO se resuelve desde snapshot
        assert float(item.unit_price) == float(sent_price)

        # Verificar que NO hay snapshots para este acopio
        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
            )
        )
        snapshots = list(result.scalars().all())
        assert len(snapshots) == 0


# ---------------------------------------------------------------------------
# Tests: Snapshot model soft-delete
# ---------------------------------------------------------------------------


class TestSnapshotModel:
    """Modelo StockpilePriceSnapshot soft-delete behavior."""

    @pytest.mark.asyncio
    async def test_soft_delete_snapshot(
        self,
        db: AsyncSession,
        user_a: User,
        business_a: Business,
    ):
        """Soft delete debe marcar deleted_at y excluir de queries."""
        from app.services.stockpile_service import StockpileService

        await _create_products(db, business_a.id)
        client = await _create_client(db, business_a.id)

        service = StockpileService(db)
        stockpile, _ = await service.create_by_amount(
            business_id=business_a.id,
            client_id=client.id,
            created_by=user_a.id,
            name="Acopio Soft Delete",
            currency="ARS",
            exchange_rate=Decimal("1.00"),
            amount=Decimal("5000.00"),
            discount_percent=Decimal("0"),
        )

        result = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
            )
        )
        snapshots = list(result.scalars().all())
        assert len(snapshots) > 0

        snapshot = snapshots[0]
        assert snapshot.deleted_at is None

        # Soft delete
        from datetime import datetime
        snapshot.deleted_at = datetime.utcnow()
        await db.commit()

        # No debe aparecer en queries con deleted_at is None
        result2 = await db.execute(
            select(StockpilePriceSnapshot).where(
                StockpilePriceSnapshot.stockpile_id == stockpile.id,
                StockpilePriceSnapshot.deleted_at.is_(None),
            )
        )
        remaining = list(result2.scalars().all())
        assert len(remaining) == len(snapshots) - 1
